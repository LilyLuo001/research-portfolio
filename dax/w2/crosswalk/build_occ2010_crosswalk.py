"""Build a fail-closed OCC2010 to O*NET-SOC 2019 crosswalk.

The primary weights come from the Census Bureau's 2010-to-2018 occupation
conversion rates.  When a 2018 Census occupation points to more than one
2018 SOC, May 2021 OEWS employment supplies the within-code allocation.
O*NET subdivisions within one detailed SOC have no official employment
counts; equal subdivision weights are emitted as provisional diagnostics and
are not marked eligible for point-dose construction.

The builder accounts for eligibility at the route component level. An earlier
implementation rejected a whole CPS occupation whenever any one SOC child had
no usable task-time shares. That made cooks, software developers, and other
large occupations disappear because of a small unavailable child. This build
keeps every officially supported component at its original Census/OEWS weight,
emits the unavailable remainder explicitly, and never renormalizes around a
gap. Mapped coverage therefore means resolved plus bounded provisional mass;
unresolved mass remains visible and cannot enter a point dose.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import pathlib
import re
from collections import defaultdict


OUTPUT_FIELDS = [
    "cps_occ2010",
    "census_2018_occ",
    "base_route_weight",
    "soc_2018_pattern",
    "soc_2018",
    "onet_soc2019",
    "mapping_weight",
    "allocation_method",
    "route_status",
    "cps_code_status",
    "downstream_eligible",
]


def sha256(path: pathlib.Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def census_code(value: object) -> str:
    if value is None or value == "":
        raise ValueError("missing Census occupation code")
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip().zfill(4)


def _workbook(path: pathlib.Path):
    try:
        import openpyxl
    except ImportError as error:
        raise RuntimeError("openpyxl is required for official crosswalk files") from error
    return openpyxl.load_workbook(path, read_only=True, data_only=True)


def read_total_conversion_rates(path: pathlib.Path) -> dict[tuple[str, str], float]:
    workbook = _workbook(path)
    try:
        sheet = workbook["E1 Total"]
        result: dict[tuple[str, str], float] = {}
        old_code = None
        for row in sheet.iter_rows(min_row=4, max_col=5, values_only=True):
            if row[0] is not None:
                old_code = census_code(row[0])
            if old_code is None or row[2] is None or row[4] is None:
                continue
            key = (old_code, census_code(row[2]))
            if key in result:
                raise ValueError(f"duplicate Census conversion-rate pair {key}")
            rate = float(row[4])
            if not 0 <= rate <= 1:
                raise ValueError(f"invalid Census conversion rate for {key}: {rate}")
            result[key] = rate
    finally:
        workbook.close()
    return result


def read_census_crosswalk(
    path: pathlib.Path,
) -> tuple[dict[str, tuple[str, str]], dict[str, str]]:
    workbook = _workbook(path)
    try:
        sheet = workbook["2010 to 2018 Crosswalk "]
        direct: dict[str, tuple[str, str]] = {}
        soc_by_new_code: dict[str, str] = {}
        for row in sheet.iter_rows(min_row=5, max_col=6, values_only=True):
            if row[3] is None or row[4] is None:
                continue
            new_code = census_code(row[4])
            soc_pattern = str(row[3]).strip().upper()
            prior_soc = soc_by_new_code.setdefault(new_code, soc_pattern)
            if prior_soc != soc_pattern:
                raise ValueError(f"conflicting SOC patterns for Census code {new_code}")
            if row[1] is not None:
                old_code = census_code(row[1])
                direct[old_code] = (new_code, soc_pattern)
    finally:
        workbook.close()
    if not direct or not soc_by_new_code:
        raise ValueError("Census crosswalk contains no mappings")
    return direct, soc_by_new_code


def read_census_titles(path: pathlib.Path) -> dict[str, str]:
    """2010 Census occupation labels for the private gap audit."""
    workbook = _workbook(path)
    result: dict[str, str] = {}
    try:
        sheet = workbook["2010 to 2018 Crosswalk "]
        for row in sheet.iter_rows(min_row=5, max_col=3, values_only=True):
            if row[1] is not None and row[2] is not None:
                result[census_code(row[1])] = str(row[2]).strip()
    finally:
        workbook.close()
    return result


def census_routes(
    crosswalk: tuple[dict[str, tuple[str, str]], dict[str, str]],
    rates: dict[tuple[str, str], float],
) -> dict[str, list[tuple[str, str, float]]]:
    direct, soc_by_new_code = crosswalk
    rates_by_old: dict[str, dict[str, float]] = defaultdict(dict)
    for (old_code, new_code), rate in rates.items():
        rates_by_old[old_code][new_code] = rate
    routes: dict[str, list[tuple[str, str, float]]] = {}
    for old_code in sorted(set(direct) | set(rates_by_old)):
        if old_code in rates_by_old:
            rate_by_new = rates_by_old[old_code]
            missing_soc = sorted(set(rate_by_new) - set(soc_by_new_code))
            if missing_soc:
                raise ValueError(
                    f"missing 2018 SOC patterns for Census codes {missing_soc}"
                )
            raw_sum = sum(rate_by_new.values())
            if abs(raw_sum - 1.0) > 0.001:
                raise ValueError(
                    f"Census conversion rates for {old_code} sum to {raw_sum:.6f}"
                )
            routes[old_code] = [
                (new_code, soc_by_new_code[new_code], rate / raw_sum)
                for new_code, rate in sorted(rate_by_new.items())
            ]
        else:
            new_code, soc_pattern = direct[old_code]
            routes[old_code] = [(new_code, soc_pattern, 1.0)]
    return routes


def read_ipums_occ2010_crosswalk(path: pathlib.Path) -> dict[str, list[str]]:
    """Read the source-owner bridge from 2018 Census OCC to IPUMS OCC2010."""
    workbook = _workbook(path)
    result: dict[str, list[str]] = defaultdict(list)
    try:
        sheet = workbook.worksheets[0]
        for row in sheet.iter_rows(min_row=2, max_col=4, values_only=True):
            if row[0] is None or row[2] is None:
                continue
            old_code = census_code(row[0])
            new_code = census_code(row[2])
            if new_code not in result[old_code]:
                result[old_code].append(new_code)
    finally:
        workbook.close()
    return dict(result)


def add_ipums_single_source_fallbacks(
    routes: dict[str, list[tuple[str, str, float]]],
    ipums_crosswalk: dict[str, list[str]],
    soc_by_new_code: dict[str, str],
) -> set[str]:
    """Add only unambiguous collapsed codes; multi-source inversions fail closed."""
    added = set()
    for old_code, new_codes in ipums_crosswalk.items():
        if old_code in routes or len(new_codes) != 1:
            continue
        new_code = new_codes[0]
        if new_code not in soc_by_new_code:
            continue
        routes[old_code] = [(new_code, soc_by_new_code[new_code], 1.0)]
        added.add(old_code)
    return added


def read_usable_onet_codes(path: pathlib.Path) -> set[str]:
    usable = set()
    with path.open(newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        required = {"onet_soc", "primary_usable"}
        missing = required - set(reader.fieldnames or [])
        if missing:
            raise ValueError(f"O*NET time shares missing fields {sorted(missing)}")
        for row in reader:
            if row["primary_usable"].strip().lower() == "true":
                usable.add(row["onet_soc"].strip())
    if not usable:
        raise ValueError("O*NET time shares contain no usable occupations")
    return usable


def read_legacy_fallback_codes(path: pathlib.Path | None) -> set[str]:
    if path is None:
        return set()
    result = set()
    with path.open(newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        if "onet_soc2019" not in (reader.fieldnames or []):
            raise ValueError("legacy fallback file missing onet_soc2019")
        for row in reader:
            result.add(row["onet_soc2019"].strip())
    return result


def read_onet_crosswalk(
    path: pathlib.Path, usable_onet_codes: set[str] | None = None
) -> dict[str, list[str]]:
    workbook = _workbook(path)
    result: dict[str, list[str]] = defaultdict(list)
    try:
        sheet = workbook.worksheets[0]
        for row in sheet.iter_rows(min_row=5, max_col=4, values_only=True):
            if row[0] is None or row[2] is None:
                continue
            onet = str(row[0]).strip()
            soc = str(row[2]).strip().upper()
            if (
                usable_onet_codes is None or onet in usable_onet_codes
            ) and onet not in result[soc]:
                result[soc].append(onet)
    finally:
        workbook.close()
    return {soc: sorted(codes) for soc, codes in result.items()}


def read_oews_employment(path: pathlib.Path) -> dict[str, float]:
    employment = {}
    with path.open(newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        required = {"vintage", "occ_code", "occupation_group", "total_employment"}
        missing = required - set(reader.fieldnames or [])
        if missing:
            raise ValueError(f"OEWS wage panel missing fields {sorted(missing)}")
        for row in reader:
            if row["vintage"] != "2021" or row["occupation_group"] != "detailed":
                continue
            value = row["total_employment"].strip()
            if value and float(value) > 0:
                employment[row["occ_code"].strip().upper()] = float(value)
    return employment


def matching_soc_codes(pattern: str, available: set[str]) -> list[str]:
    if pattern in available:
        return [pattern]
    if "X" in pattern:
        expression = "^" + re.escape(pattern).replace("X", r"\d") + "$"
    elif re.fullmatch(r"\d{2}-0000", pattern):
        expression = "^" + re.escape(pattern[:3]) + r"\d{4}$"
    elif re.fullmatch(r"\d{2}-\d{2}00", pattern):
        expression = "^" + re.escape(pattern[:-2]) + r"\d{2}$"
    elif re.fullmatch(r"\d{2}-\d{3}0", pattern):
        expression = "^" + re.escape(pattern[:-1]) + r"\d$"
    else:
        return []
    return sorted(code for code in available if re.fullmatch(expression, code))


def compose_rows(
    routes: dict[str, list[tuple[str, str, float]]],
    onet_by_soc: dict[str, list[str]],
    oews_employment: dict[str, float],
    ipums_fallback_codes: set[str] | None = None,
    all_onet_by_soc: dict[str, list[str]] | None = None,
    legacy_onet_codes: set[str] | None = None,
) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    all_available_soc = set(all_onet_by_soc or onet_by_soc)
    ipums_fallback_codes = ipums_fallback_codes or set()
    legacy_onet_codes = legacy_onet_codes or set()
    for old_code, old_routes in sorted(routes.items()):
        code_rows = []
        base_method = (
            "ipums_single_source_bridge"
            if old_code in ipums_fallback_codes
            else "census_conversion_rate"
        )
        for new_code, pattern, base_weight in old_routes:
            soc_codes = matching_soc_codes(pattern, all_available_soc)
            if not soc_codes:
                code_rows.append(
                    {
                        "cps_occ2010": old_code,
                        "census_2018_occ": new_code,
                        "base_route_weight": base_weight,
                        "soc_2018_pattern": pattern,
                        "soc_2018": "",
                        "onet_soc2019": "",
                        "mapping_weight": base_weight,
                        "allocation_method": "unresolved",
                        "route_status": "unresolved_no_onet_soc",
                    }
                )
                continue
            if len(soc_codes) == 1:
                soc_shares = {soc_codes[0]: 1.0}
                soc_method = "single_soc"
            else:
                missing_employment = [
                    code for code in soc_codes if code not in oews_employment
                ]
                if missing_employment:
                    # OEWS 2021 retains some broad/aggregate codes instead of
                    # every 2018 detailed SOC. Equal shares are only a
                    # diagnostic center; W3 must use child-dose min/max bounds.
                    soc_shares = {code: 1.0 / len(soc_codes) for code in soc_codes}
                    soc_method = "equal_soc_missing_oews"
                else:
                    employment_sum = sum(oews_employment[code] for code in soc_codes)
                    soc_shares = {
                        code: oews_employment[code] / employment_sum for code in soc_codes
                    }
                    soc_method = "oews2021_employment_share"
            for soc_code, soc_share in soc_shares.items():
                component_weight = base_weight * soc_share
                onet_codes = onet_by_soc.get(soc_code, [])
                if not onet_codes:
                    code_rows.append(
                        {
                            "cps_occ2010": old_code,
                            "census_2018_occ": new_code,
                            "base_route_weight": base_weight,
                            "soc_2018_pattern": pattern,
                            "soc_2018": soc_code,
                            "onet_soc2019": "",
                            "mapping_weight": component_weight,
                            "allocation_method": f"{base_method}+{soc_method}",
                            "route_status": "unresolved_no_usable_onet",
                        }
                    )
                    continue
                provisional = len(onet_codes) > 1
                uses_legacy = any(code in legacy_onet_codes for code in onet_codes)
                for onet_code in onet_codes:
                    code_rows.append(
                        {
                            "cps_occ2010": old_code,
                            "census_2018_occ": new_code,
                            "base_route_weight": base_weight,
                            "soc_2018_pattern": pattern,
                            "soc_2018": soc_code,
                            "onet_soc2019": onet_code,
                            "mapping_weight": component_weight / len(onet_codes),
                            "allocation_method": "+".join(filter(None, [
                                base_method,
                                soc_method,
                                "equal_within_soc" if provisional else "",
                                "onet25_legacy_bridge" if uses_legacy else "",
                            ])),
                            "route_status": (
                                "provisional_equal_soc_missing_oews"
                                if soc_method == "equal_soc_missing_oews"
                                else "provisional_legacy_task_ratings"
                                if uses_legacy
                                else "provisional_equal_within_soc"
                                if provisional
                                else "resolved_employment_weighted"
                            ),
                        }
                    )
        unresolved_weight = sum(
            float(row["mapping_weight"])
            for row in code_rows
            if str(row["route_status"]).startswith("unresolved_")
        )
        statuses = {str(row["route_status"]) for row in code_rows}
        if unresolved_weight >= 1.0 - 1e-9:
            code_status = "unresolved"
        elif unresolved_weight > 1e-9:
            code_status = "partial_unresolved"
        elif any(status.startswith("provisional_") for status in statuses):
            code_status = "provisional_equal_within_soc"
        else:
            code_status = "resolved_employment_weighted"
        for row in code_rows:
            row["cps_code_status"] = code_status
            row["downstream_eligible"] = code_status == "resolved_employment_weighted"
        rows.extend(code_rows)
    return rows


def observed_cps_codes(path: pathlib.Path | None) -> set[str]:
    if path is None:
        return set()
    result = set()
    with path.open(newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        if "cps_occ" not in (reader.fieldnames or []):
            raise ValueError("pre-period cells missing cps_occ")
        for row in reader:
            result.add(census_code(row["cps_occ"]))
    return result


def observed_cps_weight_mass(path: pathlib.Path | None) -> dict[str, float]:
    if path is None:
        return {}
    result: dict[str, float] = defaultdict(float)
    with path.open(newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        if "weight_sum" not in (reader.fieldnames or []):
            return {}
        for row in reader:
            result[census_code(row["cps_occ"])] += float(row["weight_sum"])
    return dict(result)


def component_weight_mass_shares(
    rows: list[dict[str, object]], observed_mass: dict[str, float]
) -> dict[str, float]:
    """CPS-weighted route coverage without renormalizing missing components."""
    total = sum(observed_mass.values())
    if total <= 0:
        return {}
    rows_by_code: dict[str, list[dict[str, object]]] = defaultdict(list)
    for row in rows:
        rows_by_code[str(row["cps_occ2010"])].append(row)
    result: dict[str, float] = defaultdict(float)
    for code, mass in observed_mass.items():
        code_rows = rows_by_code.get(code)
        if not code_rows:
            result["absent_from_crosswalk"] += mass / total
            continue
        for row in code_rows:
            status = str(row["route_status"])
            if status == "resolved_employment_weighted":
                bucket = "resolved"
            elif status.startswith("provisional_"):
                bucket = status
            else:
                bucket = status
            result[bucket] += mass / total * float(row["mapping_weight"])
    return dict(result)


def write_gap_audit(
    path: pathlib.Path,
    rows: list[dict[str, object]],
    observed_mass: dict[str, float],
    titles: dict[str, str],
) -> None:
    """Write the private occupation-level coverage decomposition."""
    total = sum(observed_mass.values())
    grouped: dict[str, list[dict[str, object]]] = defaultdict(list)
    for row in rows:
        grouped[str(row["cps_occ2010"])].append(row)
    fields = [
        "cps_occ2010", "occupation_title", "cps_weight_sum",
        "cps_weight_share", "cps_code_status", "resolved_component_weight",
        "provisional_component_weight", "unresolved_component_weight",
        "unresolved_reasons", "equal_split_bounds_required",
    ]
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, lineterminator="\n")
        writer.writeheader()
        for code, mass in sorted(observed_mass.items(), key=lambda item: -item[1]):
            code_rows = grouped.get(code, [])
            resolved = sum(float(r["mapping_weight"]) for r in code_rows
                           if r["route_status"] == "resolved_employment_weighted")
            provisional = sum(float(r["mapping_weight"]) for r in code_rows
                              if str(r["route_status"]).startswith("provisional_"))
            unresolved = max(0.0, 1.0 - resolved - provisional)
            reasons = sorted({str(r["route_status"]) for r in code_rows
                              if str(r["route_status"]).startswith("unresolved_")})
            writer.writerow({
                "cps_occ2010": code,
                "occupation_title": titles.get(code, ""),
                "cps_weight_sum": f"{mass:.8f}",
                "cps_weight_share": f"{(mass / total if total else 0):.12f}",
                "cps_code_status": (code_rows[0]["cps_code_status"]
                                    if code_rows else "absent_from_crosswalk"),
                "resolved_component_weight": f"{resolved:.12f}",
                "provisional_component_weight": f"{provisional:.12f}",
                "unresolved_component_weight": f"{unresolved:.12f}",
                "unresolved_reasons": "|".join(reasons),
                "equal_split_bounds_required": str(provisional > 0).lower(),
            })
    path.chmod(0o600)


def build(
    conversion_rates_path: pathlib.Path,
    census_crosswalk_path: pathlib.Path,
    onet_crosswalk_path: pathlib.Path,
    onet_timeshares_path: pathlib.Path,
    oews_wages_path: pathlib.Path,
    output_path: pathlib.Path,
    receipt_path: pathlib.Path,
    preperiod_cells_path: pathlib.Path | None = None,
    ipums_occ_crosswalk_path: pathlib.Path | None = None,
    legacy_fallback_path: pathlib.Path | None = None,
    gap_audit_path: pathlib.Path | None = None,
    coverage_threshold: float = 0.90,
    max_unresolved_occupation_share: float = 0.01,
):
    rates = read_total_conversion_rates(conversion_rates_path)
    census_crosswalk = read_census_crosswalk(census_crosswalk_path)
    routes = census_routes(census_crosswalk, rates)
    ipums_fallback_codes = set()
    if ipums_occ_crosswalk_path is not None:
        ipums_fallback_codes = add_ipums_single_source_fallbacks(
            routes,
            read_ipums_occ2010_crosswalk(ipums_occ_crosswalk_path),
            census_crosswalk[1],
        )
    usable_onet = read_usable_onet_codes(onet_timeshares_path)
    legacy_onet = read_legacy_fallback_codes(legacy_fallback_path)
    all_onet_by_soc = read_onet_crosswalk(onet_crosswalk_path)
    onet_by_soc = read_onet_crosswalk(
        onet_crosswalk_path, usable_onet | legacy_onet
    )
    rows = compose_rows(
        routes,
        onet_by_soc,
        read_oews_employment(oews_wages_path),
        ipums_fallback_codes,
        all_onet_by_soc,
        legacy_onet,
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=OUTPUT_FIELDS, lineterminator="\n")
        writer.writeheader()
        for row in rows:
            writer.writerow(
                {
                    **row,
                    "base_route_weight": f"{row['base_route_weight']:.10f}",
                    "mapping_weight": f"{row['mapping_weight']:.12f}",
                    "downstream_eligible": str(row["downstream_eligible"]).lower(),
                }
            )
    output_path.chmod(0o600)

    sums = defaultdict(float)
    statuses = {}
    for row in rows:
        sums[row["cps_occ2010"]] += float(row["mapping_weight"])
        statuses[row["cps_occ2010"]] = row["cps_code_status"]
    bad_sums = {code: value for code, value in sums.items() if abs(value - 1) > 1e-9}
    if bad_sums:
        raise ValueError(f"crosswalk weights do not sum to one: {bad_sums}")

    observed = observed_cps_codes(preperiod_cells_path)
    observed_mass = observed_cps_weight_mass(preperiod_cells_path)
    observed_missing = sorted(observed - set(statuses))
    mass_by_status: dict[str, float] = defaultdict(float)
    for code, mass in observed_mass.items():
        mass_by_status[statuses.get(code, "absent_from_crosswalk")] += mass
    total_mass = sum(mass_by_status.values())
    mass_shares = {
        status: mass / total_mass
        for status, mass in mass_by_status.items()
        if total_mass
    }
    unresolved_reasons: dict[str, set[str]] = defaultdict(set)
    for row in rows:
        if row["cps_occ2010"] in observed and row["cps_code_status"] in {
            "unresolved", "partial_unresolved"
        }:
            if str(row["route_status"]).startswith("unresolved_"):
                unresolved_reasons[row["route_status"]].add(row["cps_occ2010"])
    component_shares = component_weight_mass_shares(rows, observed_mass)
    mapped_component_share = (
        component_shares.get("resolved", 0.0)
        + sum(value for status, value in component_shares.items()
              if status.startswith("provisional_"))
    )
    total_observed_mass = sum(observed_mass.values())
    rows_by_code: dict[str, list[dict[str, object]]] = defaultdict(list)
    for row in rows:
        rows_by_code[str(row["cps_occ2010"])].append(row)
    unresolved_contributions: dict[str, float] = {}
    for code, mass in observed_mass.items():
        code_rows = rows_by_code.get(code, [])
        unresolved_weight = (
            sum(float(row["mapping_weight"]) for row in code_rows
                if str(row["route_status"]).startswith("unresolved_"))
            if code_rows else 1.0
        )
        unresolved_contributions[code] = (
            mass / total_observed_mass * unresolved_weight
            if total_observed_mass else 0.0
        )
    worst_unresolved_code = max(
        unresolved_contributions, key=unresolved_contributions.get, default=""
    )
    worst_unresolved_share = unresolved_contributions.get(worst_unresolved_code, 0.0)
    coverage_gate_pass = (
        mapped_component_share >= coverage_threshold
        and worst_unresolved_share <= max_unresolved_occupation_share
    )
    if gap_audit_path is not None:
        write_gap_audit(
            gap_audit_path, rows, observed_mass,
            read_census_titles(census_crosswalk_path),
        )
    input_paths = [
        conversion_rates_path,
        census_crosswalk_path,
        onet_crosswalk_path,
        onet_timeshares_path,
        oews_wages_path,
    ]
    if ipums_occ_crosswalk_path is not None:
        input_paths.append(ipums_occ_crosswalk_path)
    if legacy_fallback_path is not None:
        input_paths.append(legacy_fallback_path)
    receipt = {
        "status": "CPS_OCC2010_TO_ONET_SOC2019_CROSSWALK_PRIVATE",
        "weight_rule": (
            "Census total 2010-to-2018 conversion rate; May 2021 OEWS employment "
            "share across matched 2018 SOCs; unambiguous IPUMS collapsed-code "
            "bridges receive unit route weight; equal O*NET subdivision is provisional"
        ),
        "fail_closed_rule": (
            "Only cps_code_status=resolved_employment_weighted is downstream eligible"
        ),
        "inputs": {
            path.name: sha256(path)
            for path in input_paths
        },
        "output_name": output_path.name,
        "output_sha256": sha256(output_path),
        "n_mapping_rows": len(rows),
        "n_cps_codes": len(statuses),
        "n_ipums_single_source_fallback_codes": len(ipums_fallback_codes),
        "n_resolved_employment_weighted_codes": sum(
            value == "resolved_employment_weighted" for value in statuses.values()
        ),
        "n_provisional_equal_within_soc_codes": sum(
            value == "provisional_equal_within_soc" for value in statuses.values()
        ),
        "n_provisional_bounded_codes": sum(
            value == "provisional_equal_within_soc" for value in statuses.values()
        ),
        "n_unresolved_codes": sum(value == "unresolved" for value in statuses.values()),
        "n_partial_unresolved_codes": sum(
            value == "partial_unresolved" for value in statuses.values()
        ),
        "bad_weight_sums": len(bad_sums),
        "n_observed_preperiod_cps_codes": len(observed),
        "n_observed_preperiod_resolved_codes": sum(
            statuses.get(code) == "resolved_employment_weighted" for code in observed
        ),
        "n_observed_preperiod_provisional_codes": sum(
            statuses.get(code) == "provisional_equal_within_soc" for code in observed
        ),
        "n_observed_preperiod_unresolved_codes": sum(
            statuses.get(code) == "unresolved" for code in observed
        ),
        "n_observed_preperiod_partial_unresolved_codes": sum(
            statuses.get(code) == "partial_unresolved" for code in observed
        ),
        "observed_preperiod_weight_mass_shares": {
            status: round(value, 10) for status, value in sorted(mass_shares.items())
        },
        "observed_preperiod_component_weight_mass_shares": {
            status: round(value, 10)
            for status, value in sorted(component_shares.items())
        },
        "mapped_component_weight_mass_share": round(mapped_component_share, 10),
        "coverage_gate_threshold": coverage_threshold,
        "max_unresolved_occupation_weight_share_threshold": (
            max_unresolved_occupation_share
        ),
        "max_unresolved_occupation_weight_share": round(worst_unresolved_share, 10),
        "max_unresolved_occupation_code": worst_unresolved_code,
        "coverage_gate_pass": coverage_gate_pass,
        "provisional_bound_rule": (
            "For every provisional component, carry the minimum and maximum "
            "dose across its officially linked SOC/O*NET children or legacy "
            "sources; equal weights are a diagnostic center only."
        ),
        "observed_unresolved_reason_code_counts": {
            reason: len(codes) for reason, codes in sorted(unresolved_reasons.items())
        },
        "observed_codes_absent_from_official_crosswalk": observed_missing,
        "raw_workbooks_committed": False,
        "detailed_crosswalk_committed": False,
    }
    receipt_path.parent.mkdir(parents=True, exist_ok=True)
    receipt_path.write_text(json.dumps(receipt, indent=2) + "\n", encoding="utf-8")
    return receipt


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--conversion-rates", type=pathlib.Path, required=True)
    parser.add_argument("--census-crosswalk", type=pathlib.Path, required=True)
    parser.add_argument("--onet-crosswalk", type=pathlib.Path, required=True)
    parser.add_argument("--onet-timeshares", type=pathlib.Path, required=True)
    parser.add_argument("--oews-wages", type=pathlib.Path, required=True)
    parser.add_argument("--preperiod-cells", type=pathlib.Path)
    parser.add_argument("--ipums-occ-crosswalk", type=pathlib.Path)
    parser.add_argument("--legacy-fallback", type=pathlib.Path)
    parser.add_argument("--output", type=pathlib.Path, required=True)
    parser.add_argument("--receipt", type=pathlib.Path, required=True)
    parser.add_argument("--gap-audit", type=pathlib.Path)
    parser.add_argument("--coverage-threshold", type=float, default=0.90)
    parser.add_argument("--max-unresolved-occupation-share", type=float, default=0.01)
    args = parser.parse_args()
    receipt = build(
        args.conversion_rates,
        args.census_crosswalk,
        args.onet_crosswalk,
        args.onet_timeshares,
        args.oews_wages,
        args.output,
        args.receipt,
        args.preperiod_cells,
        args.ipums_occ_crosswalk,
        args.legacy_fallback,
        args.gap_audit,
        args.coverage_threshold,
        args.max_unresolved_occupation_share,
    )
    print(json.dumps(receipt, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
