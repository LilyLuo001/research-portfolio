#!/usr/bin/env python3
"""Build the frozen sibling-imputed Rule-B Eloundou exposure values.

This is a measurement-only builder.  It reads public occupation mappings,
Eloundou's published occupation scores, and OEWS employment weights.  It does
not read CPS records or any outcome.  The rule is the one signed in
``COVERAGE_RULE_PRESPEC_v1.md``: for Census occupations with at least 95%
scored component mass, replace each unscored six-digit SOC component by the
employment-weighted mean of scored detailed occupations in its SOC broad
group, then aggregate all components with the frozen target-SOC weights.
"""
from __future__ import annotations

import argparse
import hashlib
import io
import json
import pathlib
import zipfile

import numpy as np
import pandas as pd
from openpyxl import load_workbook


MEASURES = ("dv_rating_alpha", "dv_rating_beta", "dv_rating_gamma")


def sha256(path: pathlib.Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1 << 20), b""):
            digest.update(block)
    return digest.hexdigest()


def archive_xlsx(path: pathlib.Path, suffix: str) -> io.BytesIO:
    with zipfile.ZipFile(path) as archive:
        names = [name for name in archive.namelist() if name.endswith(suffix)]
        if len(names) != 1:
            raise ValueError(f"expected one {suffix} in {path}, found {names}")
        return io.BytesIO(archive.read(names[0]))


def read_excel_compat(source, sheet_name=None, skiprows=0) -> pd.DataFrame:
    """Read simple tabular workbooks without pandas' openpyxl version gate."""
    book = load_workbook(source, read_only=True, data_only=True)
    sheet = book[sheet_name] if sheet_name else book.worksheets[0]
    rows = list(sheet.iter_rows(values_only=True))
    header = [str(value).strip() if value is not None else f"Unnamed: {index}"
              for index, value in enumerate(rows[skiprows])]
    frame = pd.DataFrame(rows[skiprows + 1:], columns=header)
    for column in frame.columns:
        frame[column] = frame[column].map(lambda value: None if value is None else str(value))
    return frame


def soc18_to_census(bls_path: pathlib.Path, census_path: pathlib.Path) -> pd.DataFrame:
    xwalk = read_excel_compat(bls_path, skiprows=8)
    xwalk.columns = ["soc_2010", "soc_title_2010", "soc_2018", "soc_title_2018"]
    xwalk["soc_2018"] = xwalk["soc_2018"].str.strip()
    universe = pd.DataFrame({"soc_2018": sorted(xwalk.soc_2018.dropna().unique())})
    universe["trunc_3dig"] = universe.soc_2018.str[:4] + "XXX"
    universe["trunc_2dig"] = universe.soc_2018.str[:5] + "XX"
    universe["trunc_1dig"] = universe.soc_2018.str[:6] + "X"

    census = read_excel_compat(
        census_path, sheet_name="2018 Census Occ Code List", skiprows=4
    ).rename(columns=lambda value: str(value).strip())
    census = census.drop(columns=[c for c in census if c.startswith("Unnamed")])
    census = census.rename(columns={
        "2018 Census Title": "occupation",
        "2018 Census Code": "census_2018",
        "2018 SOC Code": "soc_marker",
    }).dropna(subset=["occupation", "census_2018", "soc_marker"])
    census = census.loc[
        ~census.occupation.str.contains(":", na=False) & census.soc_marker.ne("none")
    ].copy()
    for column in ("occupation", "census_2018", "soc_marker"):
        census[column] = census[column].str.strip()
    census["marker2"] = census.soc_marker.str.replace(r"0$", "X", regex=True)
    census["marker2"] = census.marker2.replace({
        "25-100X": "25-1XXX", "29-900X": "29-90XX",
        "39-100X": "39-10XX", "53-100X": "53-1XXX",
    })
    rows = []
    for row in census.itertuples(index=False):
        hits = []
        for column in ("trunc_1dig", "trunc_2dig", "trunc_3dig"):
            hits = universe.loc[universe[column].eq(row.marker2), "soc_2018"].tolist()
            if hits:
                break
        if not hits:
            hits = [row.soc_marker]
        rows.extend({"census_2018": row.census_2018,
                     "occupation": row.occupation, "soc_2018": soc}
                    for soc in hits)
    return pd.DataFrame(rows).drop_duplicates()


def target_employment(path: pathlib.Path) -> pd.DataFrame:
    frame = read_excel_compat(archive_xlsx(path, "national_M2021_dl.xlsx"))
    frame.columns = [column.lower() for column in frame.columns]
    frame = frame.loc[
        frame.o_group.str.lower().eq("detailed"), ["occ_code", "tot_emp"]
    ].rename(columns={"occ_code": "soc_2018"})
    frame["target_soc_employment"] = pd.to_numeric(
        frame.tot_emp.str.replace(",", "", regex=False), errors="coerce"
    )
    return frame[["soc_2018", "target_soc_employment"]]


def build(args: argparse.Namespace) -> tuple[pd.DataFrame, dict]:
    mapping = soc18_to_census(args.bls_crosswalk, args.census_crosswalk)
    employment = target_employment(args.oews_2021)
    raw = pd.read_csv(args.eloundou, dtype={"O*NET-SOC Code": str})
    raw["soc_2018"] = raw["O*NET-SOC Code"].str.strip().str[:7]
    for measure in MEASURES:
        raw[measure] = pd.to_numeric(raw[measure], errors="coerce")
    scores = raw.groupby("soc_2018", as_index=False)[list(MEASURES)].mean()
    scored = scores.merge(employment, on="soc_2018", how="left")
    scored["broad_group"] = scored.soc_2018.str[:6]

    sibling = {}
    for measure in MEASURES:
        values = {}
        for broad, group in scored.dropna(subset=[measure]).groupby("broad_group"):
            usable = group.dropna(subset=["target_soc_employment"])
            if len(usable) and usable.target_soc_employment.gt(0).all():
                values[broad] = float(np.average(
                    usable[measure], weights=usable.target_soc_employment
                ))
            else:
                values[broad] = float(group[measure].mean())
        sibling[measure] = values

    components = mapping.merge(scores, on="soc_2018", how="left").merge(
        employment, on="soc_2018", how="left"
    )
    rows = []
    for (code, title), group in components.groupby(["census_2018", "occupation"]):
        emp = group.target_soc_employment
        weights = (emp / emp.sum() if emp.notna().all() and emp.gt(0).all()
                   else pd.Series(1.0 / len(group), index=group.index))
        row = {"census2018": str(code).zfill(4), "occupation": title,
               "component_count": int(len(group))}
        for measure in MEASURES:
            available = group[measure].notna()
            covered = float(weights.loc[available].sum())
            partial = float((weights.loc[available] * group.loc[available, measure]).sum())
            completed = group[measure].copy()
            for index in group.index[~available]:
                completed.loc[index] = sibling[measure].get(group.loc[index, "soc_2018"][:6])
            row[f"{measure}_covered_mass"] = covered
            row[f"{measure}_rule_b"] = (
                float((weights * completed).sum())
                if covered >= 0.95 and completed.notna().all() else np.nan
            )
            row[f"{measure}_rule_c"] = partial / covered if covered >= 0.95 else np.nan
        rows.append(row)
    output = pd.DataFrame(rows).sort_values("census2018")

    # Frozen four-row crosswalk decomposition: deliberately naïve exact-code
    # SOC-2010-to-SOC-2018 match versus the repaired administrative mapping.
    aioe = read_excel_compat(args.aioe, sheet_name="Appendix A")
    aioe = aioe.rename(columns={"SOC Code": "soc_2010"})[["soc_2010", "AIOE"]]
    aioe["soc_2010"] = aioe.soc_2010.str.strip()
    aioe["AIOE"] = pd.to_numeric(aioe.AIOE, errors="coerce")
    exact = mapping.merge(aioe, left_on="soc_2018", right_on="soc_2010", how="left")
    exact = exact.merge(employment, on="soc_2018", how="left")
    exact_rows = []
    for code, group in exact.groupby("census_2018"):
        emp = group.target_soc_employment
        weights = (emp / emp.sum() if emp.notna().all() and emp.gt(0).all()
                   else pd.Series(1.0 / len(group), index=group.index))
        exact_rows.append({
            "census2018": str(code).zfill(4),
            "aioe_exact_code_baseline": (
                float((weights * group.AIOE).sum()) if group.AIOE.notna().all() else np.nan
            ),
        })
    output = output.merge(pd.DataFrame(exact_rows), on="census2018", how="left")

    frozen = pd.read_csv(args.frozen_variants, dtype={"census_2018": str})
    frozen["census_2018"] = frozen.census_2018.str.zfill(4)
    check = output.merge(frozen, left_on="census2018", right_on="census_2018")
    maximum_mass_gap = 0.0
    maximum_partial_gap = 0.0
    for measure in MEASURES:
        maximum_mass_gap = max(maximum_mass_gap, float(np.nanmax(np.abs(
            check[f"{measure}_covered_mass"]
            - check[f"{measure}_target_soc_covered_weight"]
        ))))
        reconstructed_partial = check[f"{measure}_rule_c"] * check[f"{measure}_covered_mass"]
        maximum_partial_gap = max(maximum_partial_gap, float(np.nanmax(np.abs(
            reconstructed_partial - check[f"{measure}_target_soc_partial_weighted_sum"]
        ))))
    if maximum_mass_gap > 1e-9 or maximum_partial_gap > 1e-9:
        raise ValueError("Rule-B component reconstruction does not match frozen variants")

    receipt = {
        "record_version": "yax-rule-b-sibling-imputation-v1",
        "status": "PASS_MEASUREMENT_ONLY",
        "post_outcomes_read": False,
        "rule": "sibling-imputed within six-digit SOC broad group; s_c >= 0.95",
        "inputs": {name: {"path": str(path), "sha256": sha256(path)} for name, path in {
            "eloundou": args.eloundou, "aioe": args.aioe,
            "bls_crosswalk": args.bls_crosswalk,
            "census_crosswalk": args.census_crosswalk, "oews_2021": args.oews_2021,
            "frozen_variants": args.frozen_variants,
        }.items()},
        "rows": int(len(output)),
        "maximum_frozen_covered_mass_gap": maximum_mass_gap,
        "maximum_frozen_partial_sum_gap": maximum_partial_gap,
    }
    return output, receipt


def main(argv=None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--eloundou", type=pathlib.Path, required=True)
    parser.add_argument("--aioe", type=pathlib.Path, required=True)
    parser.add_argument("--bls-crosswalk", type=pathlib.Path, required=True)
    parser.add_argument("--census-crosswalk", type=pathlib.Path, required=True)
    parser.add_argument("--oews-2021", type=pathlib.Path, required=True)
    parser.add_argument("--frozen-variants", type=pathlib.Path, required=True)
    parser.add_argument("--output", type=pathlib.Path, required=True)
    parser.add_argument("--receipt", type=pathlib.Path, required=True)
    args = parser.parse_args(argv)
    output, receipt = build(args)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    output.to_csv(args.output, index=False)
    receipt["output"] = {"path": str(args.output), "sha256": sha256(args.output)}
    args.receipt.write_text(json.dumps(receipt, indent=2) + "\n", encoding="utf-8")
    print(json.dumps({"status": receipt["status"], "rows": receipt["rows"]}, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
